# -*- coding: utf-8 -*-
"""
Created on Mon Nov  6 22:21:29 2017

@author: PANCH MUKESH
"""

#Make sure you  have installed Selenium
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
import datetime
import openpyxl


web = webdriver.Chrome('//Your Path//chromedriver.exe')
web.get('http://web.whatsapp.com')
time.sleep(60)


today=datetime.datetime.now()
today=today.strftime("%A")

msg='"Your Message"'


def sendmsg(friend_name,msg):    

    input_box = web.find_element(By.XPATH, '//*[@id="side"]//input')
    input_box.clear()
    input_box.click()
    input_box.send_keys(friend_name)
    
    
    x='//span[contains(text(),' +friend_name + ')]'
    print(x)
    elem = web.find_elements_by_xpath(x)
    elem[0].click()
    elem1 = web.find_elements_by_class_name('input-container')
    for i in range(1):
        elem1[0].send_keys(msg)
        web.find_element_by_class_name('compose-btn-send').click()
    
    
    return


wb=openpyxl.load_workbook('sample.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')
row_count = sheet.max_row
for i in range (3,row_count+1):
    nam=sheet.cell(row=i,column=2).value
    nam="'"+nam+"'"
    sendmsg(nam,msg)
    time.sleep(1)
    


